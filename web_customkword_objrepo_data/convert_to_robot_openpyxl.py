import openpyxl
file_excel = 'Automation Testing Report 4tc.xlsx'
workbook_test = openpyxl.load_workbook(file_excel)
worksheet_test = workbook_test.active
file_robot = open('Web_FB_messenger_EN_test_4tc.robot', "a+")

for i in range(4,120):
	check_cell = str(worksheet_test.cell(row = i, column = 1).value)
	if check_cell.isdigit() == True :


		#number of test case
		if len(str(worksheet_test.cell(row = i, column = 1).value)) == 1:
			testcase_title = "00" + str(worksheet_test.cell(row = i, column = 1).value) + "-" + str(worksheet_test.cell(row = i, column = 3).value)
		elif len(str(worksheet_test.cell(row = i, column = 1).value)) == 2:
			testcase_title = "0" + str(worksheet_test.cell(row = i, column = 1).value) + "-" + str(worksheet_test.cell(row = i, column = 3).value)
		elif len(str(worksheet_test.cell(row = i, column = 1).value)) == 3:
			testcase_title = str(worksheet_test.cell(row = i, column = 1).value) + "-" + str(worksheet_test.cell(row = i, column = 3).value)
		file_robot.write("\n" + testcase_title)

		#title test case and login function
		if str(worksheet_test.cell(row = i, column = 3).value)[:4] == "Non-":
			testcase_user = "${email}"
			testcase_user_pass = "${password}"
		elif str(worksheet_test.cell(row = i, column = 3).value)[:4] == "User" or str(worksheet_test.cell(row = i, column = 3).value)[:4] == "Prep":
			testcase_user = "${emailNonTsel}"
			testcase_user_pass = "${passwordNonTsel}"
		elif str(worksheet_test.cell(row = i, column = 3).value)[:4] == "Post":
			testcase_user = "${emailPostpaid}"
			testcase_user_pass = "${passwordPostpaid}"
		file_robot.write("\n" + "\t" + "Login_messenger" + "    " + testcase_user + "    " + testcase_user_pass)

		#step
		row_count = 0
		check_nextcell_step = str(worksheet_test.cell(row = i + 1, column = 1).value)
		while check_nextcell_step == "None" :
			if check_nextcell_step != "None" or check_nextcell_step == "end" :
				break
			row_count = row_count + 1
			check_nextcell_step = str(worksheet_test.cell(row = i + row_count, column = 1).value)

		for j in range(0, row_count):
			check_nextcell = str(worksheet_test.cell(row = i + j, column = 10).value)
			if check_nextcell != "None":
				if str(worksheet_test.cell(row = i + j, column = 10).value)[8:13] == "types":
					func = "User_input"
					params = str(worksheet_test.cell(row = i + j, column = 10).value)[15:-1]
					file_robot.write("\n" + "\t" + func + "    " + params)
				elif str(worksheet_test.cell(row = i + j, column = 10).value)[8:17] == "clicks on":
					func = "Click_carousel_button_on_specific_location"
					order = str(worksheet_test.cell(row = i + j, column = 10).value)[27:28]
					card = str(worksheet_test.cell(row = i + j, column = 10).value)[46:47]
					params = str(worksheet_test.cell(row = i + j, column = 10).value)[60:-1]
					file_robot.write("\n" + "\t" + func + "    " + order + "    " + card + "    " + params)
				elif str(worksheet_test.cell(row = i + j, column = 10).value)[3:] == "Cancel and closing" or str(worksheet_test.cell(row = i + j, column = 10).value)[4:] == "Cancel and closing":
					func = "Cancel_and_closing_session_EN"
					file_robot.write("\n" + "\t" + func)
				elif str(worksheet_test.cell(row = i + j, column = 10).value)[3:] == "Close browser" or str(worksheet_test.cell(row = i + j, column = 10).value)[4:] == "Close browser":
					func = "Closing_session"
					file_robot.write("\n" + "\t" + func)
				elif str(worksheet_test.cell(row = i + j, column = 10).value)[8:21] == "clicks button":
					func = "Click_Button_From_Response"
					order = str(worksheet_test.cell(row = i + j, column = 10).value)[37:38]
					params = str(worksheet_test.cell(row = i + j, column = 10).value)[49:-1]
					file_robot.write("\n" + "\t" + func + "    " + order + "    " + params)

				row_count_response = 0
				check_nextcell_step_respone = str(worksheet_test.cell(row = i + j + 1, column = 10).value)
				if check_nextcell_step_respone != "None" :
					row_count_response = 1
				while check_nextcell_step_respone == "None" :
					if check_nextcell_step_respone == "end" :
						break
					row_count_response = row_count_response + 1
					check_nextcell_step_respone = str(worksheet_test.cell(row = i + j + row_count_response, column = 10).value)

				for k in range(0, row_count_response):
					if str(worksheet_test.cell(row = i + j + k, column = 11).value) == "Text" :
						func = "Check_VA_response_text"
						order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
						params = str(worksheet_test.cell(row = i + j + k, column = 13).value)
						file_robot.write("\n" + "\t" + func + "    " + order + "    " + params)
					elif str(worksheet_test.cell(row = i + j + k, column = 11).value) == "Image" :
						func = "Check_VA_response_image"
						order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
						file_robot.write("\n" + "\t" + func + "    " + order)
					elif str(worksheet_test.cell(row = i + j + k, column = 11).value) == "Carousel" :
						func = "Check_VA_response_carousel_exists"
						order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
						file_robot.write("\n" + "\t" + func + "    " + order)
					elif str(worksheet_test.cell(row = i + j + k, column = 11).value) == "Validate Carousel" :
						if str(worksheet_test.cell(row = i + j + k, column = 13).value).count('\n') > 3 :
							item = str(worksheet_test.cell(row = i + j + k, column = 13).value).split('\n')
							func = "Validate_carousel_items"
							order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
							title = item[0]
							subtitle = item[1]
							button1 = item[2]
							button2 = item[3]
							button3 = item[4]
							file_robot.write("\n" + "\t" + func + "    " + order + "    " + title + "    " + subtitle + "    " + button1 + "    " + button2 + "    " + button3)
						else :
							item = str(worksheet_test.cell(row = i + j + k, column = 13).value).split('\n')
							func = "Validate_carousel_items"
							order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
							title = item[0]
							subtitle = item[1]
							button1 = item[2]
							button2 = item[3]
							#button3 = item[4]
							file_robot.write("\n" + "\t" + func + "    " + order + "    " + title + "    " + subtitle + "    " + button1 + "    " + button2)
					elif str(worksheet_test.cell(row = i + j + k, column = 11).value) == "Text with buttons" :
						if ((worksheet_test.cell(row = i + j + k, column = 13).value).encode('utf-8')).count('\n') > 3 :
							item = str(worksheet_test.cell(row = i + j + k, column = 13).value).split('\n')
							func = "Check_VA_response_text_with_buttons"
							order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
							params = item[0]
							button1 = item[1]
							button2 = item[2]
							button3 = item[3]
							file_robot.write("\n" + "\t" + func + "    " + order + "    " + params + "    " + button1 + "    " + button2 + "    " + button3)
						else :
							item = ((worksheet_test.cell(row = i + j + k, column = 13).value).encode('utf-8')).split('\n')
							func = "Check_VA_response_text_with_2buttons"
							order = str(worksheet_test.cell(row = i + j + k, column = 12).value)
							params = item[0]
							button1 = item[1]
							button2 = item[2]
							file_robot.write("\n" + "\t" + func + "    " + order + "    " + params + "    " + button1 + "    " + button2)

		file_robot.write("\n")
file_robot.close()